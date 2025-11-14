from odoo import models, fields, api, _  # pyright: ignore[reportMissingImports]
from odoo.exceptions import UserError, ValidationError  # pyright: ignore[reportMissingImports]
import base64
import io
import openpyxl  # pyright: ignore[reportMissingModuleSource]
import json
import logging

_logger = logging.getLogger(__name__)

class AccountTemplateExportWizard(models.TransientModel):
    _name = "account.template.export.wizard"
    _description = "Exportar plantilla Excel rellenada con informes MIS"

    company_id = fields.Many2one('res.company', string="Compañía", required=True, default=lambda self: self.env.company)
    mis_instance_id = fields.Many2one('mis.report.instance', string="Instancia MIS", required=True)
    template_id = fields.Many2one('account.excel.template', string="Plantilla (.xlsm)")
    generated_file = fields.Binary(string="Fichero generado", readonly=True)
    generated_filename = fields.Char(string="Nombre fichero", readonly=True)

    def _load_template_bytes(self):
        template_data = None
        if self.template_id and self.template_id.file:
            template_data = base64.b64decode(self.template_id.file)
        else:
            import os
            module_path = os.path.dirname(__file__)
            static_path = os.path.join(module_path, '..', 'static', 'template.xlsm')
            static_path = os.path.normpath(static_path)
            if os.path.exists(static_path):
                with open(static_path, 'rb') as f:
                    template_data = f.read()
        if not template_data:
            raise UserError(_("No se encontró la plantilla XLSM."))
        return template_data

    def _get_mis_values(self):
        """Obtiene los valores del informe MIS (compatible con Odoo 17)."""
        self.ensure_one()
        mis = self.mis_instance_id
        if not mis:
            raise UserError(_("Selecciona una instancia MIS."))

        try:
            _logger.info("Iniciando extracción de datos MIS para instancia: %s", mis.name)
            
            # Obtener el periodo actual de la instancia MIS
            period = mis.period_ids and mis.period_ids[0] or False
            _logger.info("Periodo seleccionado: %s", period.name if period else "Ninguno")
            if not period:
                raise UserError(_("La instancia MIS no tiene periodos definidos."))

            # Calcular el informe para el periodo
            _logger.info("Calculando matriz MIS para periodo %s", period.name)
            
            # Intentar obtener los datos del informe
            matrix = None
            try:
                # En Odoo 17, intentar primero el método actual
                if hasattr(mis, '_compute_report_data'):
                    _logger.debug("Usando método _compute_report_data()")
                    matrix = mis._compute_report_data()
                elif hasattr(mis, '_get_report_data'):
                    _logger.debug("Usando método _get_report_data()")
                    matrix = mis._get_report_data()
                # Métodos anteriores como fallback
                elif hasattr(mis, 'compute'):
                    _logger.debug("Usando método legacy compute()")
                    matrix = mis.compute()
                elif hasattr(mis, '_compute_matrix'):
                    _logger.debug("Usando método legacy _compute_matrix()")
                    matrix = mis._compute_matrix()
                
                _logger.debug("Tipo de datos recibidos: %s", type(matrix).__name__)
                
                if matrix:
                    if isinstance(matrix, dict):
                        _logger.debug("Datos recibidos como diccionario con claves: %s", list(matrix.keys()))
                    else:
                        _logger.debug("Datos recibidos como objeto de tipo: %s", type(matrix).__name__)
                
                if not matrix:
                    raise UserError(_("No se pudo calcular la matriz de resultados del informe MIS."))
                    
            except Exception as e:
                _logger.error("Error calculando matriz MIS: %s", str(e), exc_info=True)
                raise UserError(_("Error al calcular la matriz MIS: %s") % str(e))

            # Procesar los resultados del diccionario
            values = {}
            rows_processed = 0
            values_found = 0
            
            # En Odoo 17, matrix puede tener diferentes estructuras
            _logger.debug("Estructura de datos MIS recibida: %s", matrix.keys() if matrix else "None")
            
            # Determinar la estructura del informe MIS
            kpi_data = None
            if isinstance(matrix, dict):
                if 'kpi_data' in matrix:
                    kpi_data = matrix['kpi_data']
                elif 'kpi_matrix' in matrix:
                    kpi_data = matrix['kpi_matrix']
                else:
                    kpi_data = matrix  # La matriz podría ser el diccionario de KPIs directamente
            
            if not kpi_data:
                raise UserError(_("No se encontraron datos KPI en el informe MIS."))
            
            # Procesar las líneas del informe MIS
            _logger.info("KPIs disponibles en el informe MIS:")
            for kpi_id, kpi_data in kpi_data.items():
                rows_processed += 1
                # Obtenemos el nombre y descripción
                name = None
                if isinstance(kpi_data, dict):
                    name = str(kpi_data.get('name', '') or kpi_data.get('description', '') or 
                             kpi_data.get('kpi_name', '') or kpi_data.get('account_name', '')).strip()
                elif hasattr(kpi_data, 'name'):
                    name = str(kpi_data.name).strip()
                elif hasattr(kpi_data, 'description'):
                    name = str(kpi_data.description).strip()
                
                if not name:
                    _logger.warning("KPI %s no tiene nombre o descripción", kpi_id)
                    continue
                
                _logger.info("KPI %s: %s", kpi_id, name)
                
                # Generar variantes del nombre para búsqueda flexible
                name_variants = set()
                # Versión original
                name_variants.add(name.strip())
                # Versión en minúsculas
                name_variants.add(name.lower())
                # Sin puntos
                name_variants.add(name.lower().replace('.', ''))
                # Sin espacios
                name_variants.add(name.lower().replace(' ', ''))
                # Solo alfanuméricos
                name_variants.add(''.join(c for c in name.lower() if c.isalnum()))
                # Con guiones
                name_variants.add(name.lower().replace(' ', '-'))
                # Con guiones bajos
                name_variants.add(name.lower().replace(' ', '_'))
                
                _logger.debug("Procesando KPI %s con variantes: %s", name, name_variants)
                
                try:
                    # Intentar obtener el valor según la estructura
                    val = 0.0
                    
                    if isinstance(kpi_data, dict):
                        # Estructura nueva de Odoo 17
                        if 'periods' in kpi_data:
                            period_data = kpi_data.get('periods', {}).get(str(period.id), {})
                            if isinstance(period_data, dict):
                                val = float(period_data.get('value', 0.0) or 0.0)
                            elif isinstance(period_data, (int, float)):
                                val = float(period_data)
                        # Estructura directa
                        elif 'value' in kpi_data:
                            val = float(kpi_data.get('value', 0.0) or 0.0)
                        elif 'amount' in kpi_data:
                            val = float(kpi_data.get('amount', 0.0) or 0.0)
                    elif hasattr(kpi_data, 'value'):
                        val = float(kpi_data.value or 0.0)
                    elif hasattr(kpi_data, 'amount'):
                        val = float(kpi_data.amount or 0.0)
                    elif isinstance(kpi_data, (int, float)):
                        val = float(kpi_data)
                    
                    if val != 0.0:
                        values_found += 1
                        # Guardar el valor con todas las variantes del nombre
                        for variant in name_variants:
                            if variant:  # Asegurarse de que la variante no está vacía
                                values[variant] = val
                        _logger.debug("Valor extraído para %s: %f", name, val)
                    
                except (ValueError, TypeError) as e:
                    _logger.warning("Error convirtiendo valor para %s: %s. Usando 0.0", name, str(e))
                    # Guardar 0.0 para todas las variantes
                    for variant in name_variants:
                        if variant:
                            values[variant] = 0.0

            _logger.info("Procesamiento completado: %d filas procesadas, %d valores encontrados", 
                        rows_processed, values_found)

            if not values:
                raise UserError(_("No se encontraron valores en el informe MIS."))

            # Log de algunos valores de ejemplo
            sample_values = dict(list(values.items())[:3])
            _logger.info("Muestra de valores obtenidos: %s", sample_values)

            return values

        except Exception as e:
            _logger.error("Error al obtener valores MIS: %s", str(e), exc_info=True)
            raise UserError(_("Error al calcular el MIS: %s") % str(e))

    def action_generate(self):
        self.ensure_one()
        _logger.info("Iniciando generación de plantilla Excel para compañía: %s", self.company_id.name)
        
        # Obtener valores MIS
        _logger.info("Obteniendo valores del informe MIS: %s", self.mis_instance_id.name)
        mis_values = self._get_mis_values()
        if not mis_values:
            raise UserError(_("No se encontraron valores en la instancia MIS seleccionada."))
        _logger.info("Obtenidos %d valores del informe MIS", len(mis_values))

        # Cargar plantilla
        _logger.info("Cargando plantilla Excel")
        template_data = self._load_template_bytes()
        wb = openpyxl.load_workbook(io.BytesIO(template_data), keep_vba=True)
        if 'Cuentas' not in wb.sheetnames:
            raise UserError(_("La plantilla debe tener una hoja llamada 'Cuentas'."))
        ws = wb['Cuentas']
        _logger.info("Plantilla cargada correctamente")

        # Procesar filas
        row = 14  # Comenzar desde la fila 14
        matches_found = 0
        total_rows = 0
        while True:
            cell_value = ws[f"A{row}"].value
            if not cell_value:
                _logger.info("Fin de datos en fila %d", row)
                break
                
            total_rows += 1
            # Obtener el nombre original de la celda
            original_name = str(cell_value).strip()
            
            # Generar variantes del nombre para búsqueda flexible
            name_variants = [
                original_name.lower(),  # Versión en minúsculas
                original_name.lower().replace('.', ''),  # Sin puntos
                original_name.lower().replace(' ', ''),  # Sin espacios
                ''.join(c for c in original_name.lower() if c.isalnum()),  # Solo alfanuméricos
                original_name.lower().replace(' ', '-'),  # Con guiones
                original_name.lower().replace(' ', '_')   # Con guiones bajos
            ]
            
            # Intentar encontrar el valor con cualquiera de las variantes
            val = 0.0
            found_variant = None
            for variant in name_variants:
                if variant in mis_values:
                    val = mis_values[variant]
                    found_variant = variant
                    break
            
            # Registrar si se encontró coincidencia
            if val != 0.0:
                matches_found += 1
                _logger.info("Coincidencia encontrada para '%s' usando variante '%s': %f", 
                          original_name, found_variant, val)
            else:
                _logger.warning("No se encontró coincidencia para '%s'. Variantes probadas: %s", 
                             original_name, name_variants)
            
            # Escribir el valor en la columna C, redondeado a 2 decimales
            ws[f"C{row}"] = round(val, 2)
            row += 1
            
        _logger.info("Procesamiento completado: %d coincidencias encontradas de %d filas procesadas", 
                    matches_found, total_rows)
        
        if matches_found == 0:
            _logger.warning("¡ADVERTENCIA! No se encontraron coincidencias para ninguna cuenta")

        # Si el porcentaje de coincidencias es muy bajo, mostrar una advertencia
        match_percentage = (matches_found / total_rows * 100) if total_rows > 0 else 0
        if match_percentage < 50:
            _logger.warning("Solo se encontraron coincidencias para el %.2f%% de las cuentas", 
                         match_percentage)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        file_data = out.read()
        
        # Guardar el archivo generado
        filename = f"{self.company_id.name.replace(' ', '_')}_MIS_{fields.Date.today()}"
        if self.template_id:
            filename += f"_{self.template_id.name.replace(' ', '_')}"
        filename += ".xlsm"
        
        self.write({
            'generated_file': base64.b64encode(file_data),
            'generated_filename': filename
        })

        # Si el porcentaje de coincidencias es bajo, mostrar un mensaje al usuario
        if match_percentage < 50:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Advertencia'),
                    'message': _(
                        'Solo se encontraron coincidencias para el %.2f%% de las cuentas. '
                        'Verifica que los nombres de las cuentas en la plantilla coincidan '
                        'con los del informe MIS.'
                    ) % match_percentage,
                    'type': 'warning',
                    'sticky': True,
                }
            }

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'account.template.export.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',
            'context': self._context,
        }
